#读取文件中的域名并查询相关的域名信息
#***************************************
import openpyxl
import requests
import json
from openpyxl.workbook import Workbook

webhookUrl='https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=b0913301-02c8-4f07-be0c-37380ea7828b'#调试
key='b0913301-02c8-4f07-be0c-37380ea7828b'

#发送文件到企业微信
def wx_post():
    #上传文件
    id_url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key=b0913301-02c8-4f07-be0c-37380ea7828b&type=file'  # 上传文件接口地址
    data = {'file': open('domain.xlsx', 'rb')}
    response = requests.post(url=id_url, files=data)
    json_res = response.json()
    #发送文件
    media_id = json_res['media_id']  # 提取返回ID
    wx_url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=b0913301-02c8-4f07-be0c-37380ea7828b'  # 发送消息接口地址
    data = {"msgtype": "file", "file": {"media_id": media_id}}
    r = requests.post(url=wx_url, json=data)
    return r

#读取域名文件
def open_txt():
    domains=[]
    file=open('domaintest.txt','r',encoding='utf-8')
    file_data=file.readlines()
    for row in file_data:
        tmp_list=row.split('.')
        domains.append(tmp_list[0])
    return domains

#调用搜索接口
def url_get(domain):
    # 设置请求头，模拟浏览器访问
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36"
    }

    #搜索关键词
    keyword=domain

    #发送get请求
    url = "https://api.knet.cn/whois"
    parmas={"domain":keyword}
    response=requests.get(url=url,headers=headers,params=parmas)
    res=response.text
    return res

#解析接口返回的json数据
def json_parse(res,domain):
    data=json.loads(res)
    # print(data)
    creation_data = ''
    registry_expiry_date = ''
    #判断域名是否注册
    status=data["msg"]
    if status == 'match':
        domain_match = '已注册'
        creation_data=data["data"]["creation_date"]
        # print(creation_data)
        registry_expiry_date=data["data"]["registry_expiry_date"]
        # print(registry_expiry_date)
    else:
        domain_match='无注册信息'
    return domain_match,creation_data,registry_expiry_date

#将域名信息整理成列表
def list_domain(domain,domain_match,creation_data,registry_expiry_date):
    if domain_match == '已注册':
        match_info.append(domain)
        match_info.append(domain_match)
        match_info.append(creation_data)
        match_info.append(registry_expiry_date)
        match_infos.append(match_info)
    else:
        nomatch_info.append(domain)
        nomatch_info.append(domain_match)
        nomatch_info.append(creation_data)
        nomatch_info.append(registry_expiry_date)
        nomatch_infos.append(nomatch_info)
    domain_infos=[match_infos,nomatch_infos]
    return match_infos,nomatch_infos

#将域名信息存入表格
def save_domian(match_infos,nomatch_infos):
    wb=Workbook()
    ws=wb.active
    ws.title=u'已注册信息列表'
    print(match_infos)
    print(nomatch_infos)
    r=1
    for line in match_infos:
        for col in range(1, len(line) + 1):
            ws.cell(row=r, column=col).value = line[col - 1]
        r += 1
    ws2=wb.create_sheet('无注册信息列表')
    q=1
    for line in nomatch_infos:
        for col in range(1, len(line) + 1):
            ws2.cell(row=q, column=col).value = line[col - 1]
        q += 1
    wb.save('domain.xlsx')
    wb.close()

if __name__ == '__main__':
    domains=open_txt()
    match_info=[]
    match_infos=[]
    match_infos.append(['域名','域名状态','注册时间','到期时间'])
    nomatch_info = []
    nomatch_infos = []
    nomatch_infos.append(['域名', '域名状态', '注册时间', '到期时间'])
    for domain in domains:
        res=url_get(domain)
        domain_match,creation_data,registry_expiry_date=json_parse(res,domain)
        match_infos,nomatch_infos=list_domain(domain,domain_match,creation_data,registry_expiry_date)
        match_info = []
        nomatch_info = []
    save_domian(match_infos,nomatch_infos)
    # wx_post()