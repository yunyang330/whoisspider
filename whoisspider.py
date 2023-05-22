#读取文件中的域名并查询相关的域名信息
#***************************************
import time
from time import sleep
import openpyxl
import requests
import json
from openpyxl.workbook import Workbook
import multiprocessing

max_retries=999999 #最大重试次数
retry_interval=1 #失败重试间隔
timeout=10 #请求超时

webhookUrl='https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=b0913301-02c8-4f07-be0c-37380ea7828b'#调试
key='b0913301-02c8-4f07-be0c-37380ea7828b'#调试

#发送文件到企业微信
def wx_post(key):
    #上传文件
    id_url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key='+key+'&type=file'  # 上传文件接口地址
    data = {'file': open('domain.xlsx', 'rb')}
    response = requests.post(url=id_url, files=data)
    json_res = response.json()
    #发送文件
    media_id = json_res['media_id']  # 提取返回ID
    wx_url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key='+key  # 发送消息接口地址
    data = {"msgtype": "file", "file": {"media_id": media_id}}
    r = requests.post(url=wx_url, json=data)
    return r

#读取域名文件
def open_txt():
    domains=[]
    # file=open('domaintest.txt','r',encoding='utf-8')
    file=open('domain.txt','r',encoding='utf-8')
    file_data=file.readlines()
    for row in file_data:
        domains.append(row)
    print(domains)
    return domains

#调用搜索接口
def url_get(domain):
    # 使用代理ip
    proxy_host = 'http-dynamic.xiaoxiangdaili.com'
    proxy_port = 10030
    proxy_username = '973422222989742080'
    proxy_pwd = 'PKH56rd5'

    proxyMeta = "http://%(user)s:%(pass)s@%(host)s:%(port)s" % {
        "host": proxy_host,
        "port": proxy_port,
        "user": proxy_username,
        "pass": proxy_pwd,
    }


    # 设置请求头，模拟浏览器访问
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36",
        "Connection":"close",
        'Accept-Encoding': 'gzip'
    }

    #搜索关键词
    keyword=domain

    #发送get请求
    url = "http://123.4.cn/api/main"
    parmas={"domain":keyword}
    retry_count=0
    while True:
        try:
            proxies = {
                'http': proxyMeta,
                'https': proxyMeta,
            }
            response = requests.get(url=url, headers=headers, params=parmas,proxies=proxies,timeout=timeout)
            if response.status_code == 200:
                break
            else:
                retry_count += 1
                if retry_count > max_retries:
                    raise Exception("接口访问超过最大重试次数")
        except Exception as e:
            retry_count += 1
            if retry_count > max_retries:
                raise Exception("接口访问超过最大重试次数")
            else:
                # print(f"请求接口出错，错误信息为：{e}")
                time.sleep(retry_interval)  # 等待重试间隔后再次尝试请求

    res=response.text
    response.keep_alive=False
    return res

#解析接口返回的json数据
def json_parse(res,domain):
    data=json.loads(res)
    print(data)
    creation_data = ''
    expire_date = ''
    owner_name = ''
    domain_match='域名格式不正确'
    #判断域名是否注册
    if data["retcode"]==0:
        owner_name=data["data"]["owner_name"]
        # print(owner_name)
        if owner_name != '':
            domain_match = '已注册'
            creation_data=data["data"]["create_date"]
            # print(creation_data)
            expire_date=data["data"]["expire_date"]
            # print(registry_expiry_date)
        else:
            domain_match='无注册信息'
    return owner_name,domain_match,creation_data,expire_date

#将域名信息存入表格
def save_domian(match_infos,nomatch_infos):
    wb=Workbook()
    ws=wb.active
    ws.title=u'已注册信息列表'
    r=1
    for line in match_infos:
        for col in range(1, len(line) + 1):
            ws.cell(row=r, column=col).value = line[col - 1]
        r += 1
    ws2=wb.create_sheet('无注册信息列表')
    q=1
    for line in nomatch_infos:
        for col in range(1, len(line) + 1):
            ws2.cell(row=q, column=col).value = line[col - 1]
        q += 1
    wb.save('domain.xlsx')
    wb.close()

#爬虫主程序
def main(domain):
    res = url_get(domain)
    owner_name,domain_match, creation_data, registry_expiry_date = json_parse(res, domain)
    #将域名信息整理成列表
    match_info = []
    nomatch_info = []
    if domain_match == '已注册':
        match_info.append(domain)
        match_info.append(owner_name)
        match_info.append(domain_match)
        match_info.append(creation_data)
        match_info.append(registry_expiry_date)
    # 存放无注册信息
    else:
        nomatch_info.append(domain)
        nomatch_info.append(domain_match)
        nomatch_info.append(creation_data)
        nomatch_info.append(registry_expiry_date)
    print("已查询域名：{}".format(domain))
    return match_info,nomatch_info

if __name__ == '__main__':
    multiprocessing.freeze_support()
    match_infos = []
    nomatch_infos = []
    domains=open_txt()
    #多进程运行
    with multiprocessing.Pool(processes=60) as pool:
        results = pool.map(main,domains)
        result1 = [r[0] for r in results]
        result2 = [r[1] for r in results]
        #删除列表中的[]
        for match_info in result1:
            if match_info != []:
                match_infos.append(match_info)
        for nomatch_info in result2:
            if nomatch_info != []:
                nomatch_infos.append(nomatch_info)
    pool.close()
    pool.join()
    save_domian(match_infos,nomatch_infos)
    # wx_post(key)